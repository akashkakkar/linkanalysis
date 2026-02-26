#!/usr/bin/env python3
"""Tests for Link Analyzer (regex/keyword version). Run: python -m pytest test_analyze.py -v"""

import os
import tempfile
import pytest
from openpyxl import load_workbook
from analyze import parse_date, extract_links, categorize_url, is_claude_related, get_relevance_score, get_claude_accuracy_note, analyze, generate_excel

SAMPLE_CHAT = """2/1/2026, 08:15 - Akash: Check this out https://www.linkedin.com/posts/anthropic_claude-code-tips-activity-123456
2/1/2026, 09:30 - Ravi: https://www.linkedin.com/posts/johndoe_ai-agents-activity-789012
2/1/2026, 10:00 - Akash: Great article https://www.linkedin.com/posts/janedoe_voice-tts-elevenlabs-activity-345678
3/1/2026, 11:45 - Priya: https://www.youtube.com/watch?v=abc123
3/1/2026, 14:20 - Akash: https://www.linkedin.com/posts/techguy_openclaw-claude-community-activity-456789
4/1/2026, 08:00 - Ravi: https://www.linkedin.com/posts/someone_deepseek-qwen-activity-112233
4/1/2026, 09:15 - Akash: https://www.linkedin.com/posts/dev_cursor-vscode-coding-activity-223344
5/1/2026, 10:30 - Priya: https://www.linkedin.com/posts/founder_startup-saas-revenue-activity-334455
5/1/2026, 12:00 - Akash: https://www.linkedin.com/posts/pm_productmanagement-roadmap-activity-445566
5/1/2026, 16:00 - Ravi: https://www.facebook.com/reel/some-video
6/1/2026, 07:00 - Akash: https://www.linkedin.com/posts/aidev_cowork-anthropic-activity-556677
6/1/2026, 08:30 - Priya: https://www.linkedin.com/posts/scholar_stanford-course-tutorial-activity-667788
6/1/2026, 11:00 - Akash: https://www.linkedin.com/posts/hr_hiring-remote-interview-activity-778899
7/1/2026, 09:00 - Ravi: https://www.linkedin.com/posts/coder_moltbot-automation-activity-889900
7/1/2026, 14:00 - Akash: https://mahadiscom.in/bill-payment
7/1/2026, 15:30 - Priya: https://www.linkedin.com/posts/datascientist_rag-vectordb-embedding-activity-990011
"""

class TestParseDate:
    def test_dd_mm_yyyy(self): assert parse_date("02/01/2026") == "2026-01-02"
    def test_dd_mm_yy(self): assert parse_date("02/01/26") == "2026-01-02"
    def test_yyyy_mm_dd(self): assert parse_date("2026-01-02") == "2026-01-02"
    def test_with_trailing_comma(self): assert parse_date("02/01/2026,") == "2026-01-02"
    def test_with_brackets(self): assert parse_date("[02/01/2026]") == "2026-01-02"
    def test_unparseable(self): assert parse_date("not-a-date") == "not-a-date"

class TestExtractLinks:
    def test_count(self): assert len(extract_links(SAMPLE_CHAT)) == 16
    def test_dates(self): assert extract_links(SAMPLE_CHAT)[0]["date"] == "2026-01-02"
    def test_urls_valid(self):
        for link in extract_links(SAMPLE_CHAT): assert link["url"].startswith("http")
    def test_no_trailing_punct(self):
        assert extract_links("2/1/2026, 08:00 - U: https://x.com/p.")[0]["url"] == "https://x.com/p"
    def test_empty(self): assert extract_links("") == []
    def test_no_urls(self): assert extract_links("2/1/2026, 08:00 - U: Hello!") == []
    def test_bracket_format(self):
        links = extract_links("[02/01/2026, 08:15:30] U: https://x.com/a")
        assert len(links) == 1 and links[0]["date"] == "2026-01-02"
    def test_multi_url(self): assert len(extract_links("2/1/2026, 08:00 - U: https://a.com https://b.com")) == 2
    def test_unknown_date(self): assert extract_links("https://x.com")[0]["date"] == "Unknown"

class TestCategorizeUrl:
    def test_claude(self): assert categorize_url("https://linkedin.com/posts/u_claude-code-activity-1") == "Claude/Anthropic"
    def test_anthropic(self): assert categorize_url("https://linkedin.com/posts/u_anthropic-release-activity-1") == "Claude/Anthropic"
    def test_cowork(self): assert categorize_url("https://linkedin.com/posts/u_cowork-anthropic-activity-1") == "Claude/Anthropic"
    def test_openclaw(self): assert categorize_url("https://linkedin.com/posts/u_openclaw-tool-activity-1") == "Claude/Anthropic"
    def test_voice(self): assert categorize_url("https://linkedin.com/posts/u_voice-tts-elevenlabs-activity-1") == "Voice AI / TTS"
    def test_youtube(self): assert categorize_url("https://www.youtube.com/watch?v=abc") == "YouTube"
    def test_facebook(self): assert categorize_url("https://www.facebook.com/reel/1") == "Facebook"
    def test_coding_not_claude(self): assert categorize_url("https://linkedin.com/posts/u_cursor-vscode-coding-activity-1") == "Coding / Dev Tools"
    def test_claude_code_is_claude(self): assert categorize_url("https://linkedin.com/posts/u_claude-code-tips-activity-1") == "Claude/Anthropic"
    def test_deepseek(self): assert categorize_url("https://linkedin.com/posts/u_deepseek-qwen-activity-1") == "Tech Companies"
    def test_business(self): assert categorize_url("https://linkedin.com/posts/u_startup-saas-revenue-activity-1") == "Business / Sales"
    def test_pm(self): assert categorize_url("https://linkedin.com/posts/u_productmanagement-roadmap-activity-1") == "Product Management"
    def test_education(self): assert categorize_url("https://linkedin.com/posts/u_stanford-course-activity-1") == "Education / Learning"
    def test_career(self): assert categorize_url("https://linkedin.com/posts/u_hiring-remote-interview-activity-1") == "Career / Jobs"
    def test_rag(self): assert categorize_url("https://linkedin.com/posts/u_rag-vectordb-embedding-activity-1") == "RAG / Retrieval"
    def test_generic_linkedin(self): assert categorize_url("https://linkedin.com/posts/u_vacation-photos-activity-1") == "LinkedIn Post - General"
    def test_moltbot(self): assert categorize_url("https://linkedin.com/posts/u_moltbot-automation-activity-1") == "Claude/Anthropic"
    def test_other(self): assert categorize_url("https://mahadiscom.in/bill-payment") == "Other"

class TestIsClaudeRelated:
    def test_claude_cat(self): assert is_claude_related("https://x.com", "Claude/Anthropic") is True
    def test_non_claude(self): assert is_claude_related("https://x.com/something", "Coding / Dev Tools") is False
    def test_keyword_in_url(self): assert is_claude_related("https://x.com/claude-tips", "Other") is True

class TestRelevanceScore:
    def test_claude_5(self): assert get_relevance_score("Claude/Anthropic", "https://x.com") == 5
    def test_ai_4(self): assert get_relevance_score("AI / ML General", "https://x.com") == 4
    def test_biz_3(self): assert get_relevance_score("Business / Sales", "https://x.com") == 3
    def test_yt_2(self): assert get_relevance_score("YouTube", "https://youtube.com") == 2
    def test_bill_1(self): assert get_relevance_score("Other", "https://mahadiscom.in/bill-payment") == 1
    def test_referral_1(self): assert get_relevance_score("Other", "https://porter.in/referral/abc") == 1

class TestClaudeAccuracyNote:
    def test_cowork(self): assert "Anthropic product" in get_claude_accuracy_note("https://x.com/cowork-tips")
    def test_openclaw(self): assert "NOT official" in get_claude_accuracy_note("https://x.com/openclaw-tool")
    def test_moltbot(self): assert "Third-party" in get_claude_accuracy_note("https://x.com/moltbot-tool")
    def test_claude_code(self): assert "Official" in get_claude_accuracy_note("https://x.com/claudecode-tips")
    def test_default(self): assert "docs.anthropic.com" in get_claude_accuracy_note("https://x.com/claude-general")

class TestPipeline:
    def test_full(self): assert len(analyze(SAMPLE_CHAT)) == 16
    def test_fields(self):
        for r in analyze(SAMPLE_CHAT):
            assert set(r.keys()) == {"date", "url", "category", "claude", "relevance", "accuracy_note"}
    def test_claude_detected(self): assert len([r for r in analyze(SAMPLE_CHAT) if r["claude"]]) >= 4
    def test_relevance_range(self):
        for r in analyze(SAMPLE_CHAT): assert 1 <= r["relevance"] <= 5
    def test_notes_only_claude(self):
        for r in analyze(SAMPLE_CHAT):
            if not r["claude"]: assert r["accuracy_note"] == ""
            else: assert r["accuracy_note"] != ""

class TestExcel:
    @pytest.fixture
    def excel_path(self):
        results = analyze(SAMPLE_CHAT)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        generate_excel(results, path)
        yield path
        os.unlink(path)

    def test_created(self, excel_path): assert os.path.getsize(excel_path) > 0
    def test_three_sheets(self, excel_path): assert load_workbook(excel_path).sheetnames == ["All Links", "Claude Topics", "Summary"]
    def test_row_count(self, excel_path): assert load_workbook(excel_path)["All Links"].max_row == 17
    def test_headers(self, excel_path):
        ws = load_workbook(excel_path)["All Links"]
        assert [ws.cell(1, c).value for c in range(1, 7)] == ["S.No.", "Date", "Link", "Topic/Category", "Claude Related", "Relevance (1-5)"]
    def test_claude_sheet(self, excel_path): assert load_workbook(excel_path)["Claude Topics"].max_row > 1
    def test_summary(self, excel_path):
        ws = load_workbook(excel_path)["Summary"]
        assert ws.cell(2, 1).value == "Total Links" and ws.cell(2, 2).value == 16
    def test_freeze(self, excel_path): assert load_workbook(excel_path)["All Links"].freeze_panes == "A2"
    def test_empty_ok(self):
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        generate_excel([], path)
        assert len(load_workbook(path).sheetnames) == 3
        os.unlink(path)

class TestEdgeCases:
    def test_empty(self): assert analyze("") == []
    def test_no_urls(self): assert analyze("Just text") == []
    def test_single(self):
        r = analyze("2/1/2026, 08:00 - U: https://linkedin.com/posts/u_claude-tips-activity-1")
        assert len(r) == 1 and r[0]["claude"] is True
    def test_unicode(self): assert len(analyze("2/1/2026, 08:00 - U: \U0001f680 https://x.com/ai")) == 1

if __name__ == "__main__":
    pytest.main([__file__, "-v"])
