#!/usr/bin/env python3
"""
Link Analyzer — Parse WhatsApp/chat exports, categorize URLs, and generate Excel reports.
(Regex/keyword-matching version — no API key needed, runs offline)

Usage:
    python analyze_regex.py <chat_file.txt> [output.xlsx]
    cat chat.txt | python analyze_regex.py - [output.xlsx]
"""

import re
import sys
import os
from datetime import datetime
from collections import Counter
from urllib.parse import urlparse

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

CLAUDE_KEYWORDS = [
    "claude", "anthropic", "cowork", "openclaw", "clawdbot", "zeroclaw",
    "nanoclaw", "picoclaw", "moltbot", "moltbook", "nanobots", "nanobot",
]

CATEGORY_RULES = [
    ("Claude/Anthropic", CLAUDE_KEYWORDS, []),
    ("Voice AI / TTS", ["voice", "tts", "speech", "elevenlabs", "voxcpm", "sopranotts", "text-to-speech", "whisper"], []),
    ("AI Prompting", ["prompt", "prompting", "promptengineering", "systemprompt"], []),
    ("RAG / Retrieval", ["rag", "retrieval", "vectordb", "embedding", "pinecone", "chroma", "weaviate"], []),
    ("Coding / Dev Tools", ["code", "coding", "github", "cursor", "developer", "vscode", "programming", "api", "sdk", "devtools"], CLAUDE_KEYWORDS),
    ("AI Agents", ["agent", "agentic", "crewai", "autogen", "langchain", "langgraph", "swarm"], []),
    ("Education / Learning", ["stanford", "mit", "harvard", "course", "tutorial", "learn", "mooc", "certification", "university"], []),
    ("Product Management", ["productmanagement", "productmanager", "roadmap"], []),
    ("Business / Sales", ["sales", "marketing", "sdr", "startup", "founder", "revenue", "growth", "b2b", "saas"], []),
    ("Career / Jobs", ["hiring", "remote", "interview", "layoff", "resume", "career", "job", "recruit"], []),
    ("Tech Companies", ["deepseek", "qwen", "openai", "google", "microsoft", "apple", "nvidia", "meta", "amazon", "tesla"], []),
    ("Open Source", ["opensource", "open-source", "foss", "linux", "huggingface"], []),
    ("Health / Wellness", ["health", "fitness", "medic", "yoga", "mental", "diet", "nutrition"], []),
    ("Finance", ["finance", "invest", "stock", "crypto", "trading", "money", "tax"], []),
    ("India / Culture", ["sanatan", "india", "bharati", "hindu", "diwali", "modi"], []),
    ("AI / ML General", ["ai", "llm", "genai", "machinelearning", "deeplearning", "gpt", "transformer", "diffusion", "ocr", "nlp", "chatbot"], CLAUDE_KEYWORDS),
    ("Robotics", ["robot", "robotics", "drone", "autonomous"], []),
]

PLATFORM_OVERRIDES = {
    "youtube.com": "YouTube", "youtu.be": "YouTube",
    "facebook.com": "Facebook", "fb.watch": "Facebook",
    "instagram.com": "Instagram",
    "twitter.com": "Twitter/X", "x.com": "Twitter/X",
}

RELEVANCE_SCORES = {
    "Claude/Anthropic": 5, "AI Agents": 4, "AI Prompting": 4, "RAG / Retrieval": 4,
    "Voice AI / TTS": 4, "Coding / Dev Tools": 4, "Education / Learning": 4,
    "Product Management": 4, "AI / ML General": 4, "Robotics": 4,
    "Open Source": 3, "Tech Companies": 3, "Business / Sales": 3,
    "Career / Jobs": 3, "Finance": 3, "Health / Wellness": 3,
    "India / Culture": 2, "YouTube": 2, "Facebook": 2, "Instagram": 2,
    "Twitter/X": 2, "Other": 2,
}

CLAUDE_ACCURACY_NOTES = {
    "cowork": "Real Anthropic product (beta). Verify specific feature claims against docs.anthropic.com.",
    "openclaw": "Community open-source project, NOT official Anthropic product.",
    "clawdbot": "Community open-source project, NOT official Anthropic product.",
    "zeroclaw": "Community open-source project, NOT official Anthropic product.",
    "nanoclaw": "Community tool, NOT official Anthropic product.",
    "picoclaw": "Community tool, NOT official Anthropic product.",
    "moltbot": "Third-party tool, NOT an official Anthropic product.",
    "moltbook": "Third-party tool, NOT an official Anthropic product.",
    "claude code": "Official Anthropic CLI tool. Verify version/feature claims against releases.",
    "opus": "Official model. Verify version numbers against docs.anthropic.com.",
    "sonnet": "Official model. Verify version numbers against docs.anthropic.com.",
    "haiku": "Official model. Verify version numbers against docs.anthropic.com.",
    "default": "Verify claims against docs.anthropic.com.",
}


# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------

DATE_PATTERNS = [
    r"(\d{1,2}/\d{1,2}/\d{2,4}),?\s*\d{1,2}:\d{2}",
    r"(\d{4}-\d{2}-\d{2})\s*\d{1,2}:\d{2}",
    r"(\d{1,2}/\d{1,2}/\d{4})",
    r"\[(\d{1,2}/\d{1,2}/\d{2,4}),?\s*\d{1,2}:\d{2}:\d{2}\]",
]

URL_PATTERN = re.compile(r'https?://[^\s<>"\')\]]+')


def parse_date(date_str: str) -> str:
    date_str = date_str.strip().strip("[]")
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%m/%d/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(date_str.rstrip(","), fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return date_str


def extract_links(text: str) -> list[dict]:
    results = []
    current_date = None
    for line in text.splitlines():
        for pattern in DATE_PATTERNS:
            match = re.search(pattern, line)
            if match:
                current_date = parse_date(match.group(1))
                break
        for url in URL_PATTERN.findall(line):
            url = url.rstrip(".,;:!?)]}>")
            results.append({"date": current_date or "Unknown", "url": url})
    return results


# ---------------------------------------------------------------------------
# Categorization (keyword-based)
# ---------------------------------------------------------------------------

def categorize_url(url: str) -> str:
    parsed = urlparse(url)
    domain = parsed.netloc.lower()
    path = parsed.path.lower() + (parsed.query or "").lower()
    full = (domain + path).replace("-", "").replace("_", "")

    for platform_domain, category in PLATFORM_OVERRIDES.items():
        if platform_domain in domain:
            return category

    for category, keywords, excludes in CATEGORY_RULES:
        if any(ex in full for ex in excludes):
            continue
        if any(kw in full for kw in keywords):
            return category

    if "linkedin.com" in domain:
        return "LinkedIn Post - General"
    return "Other"


def is_claude_related(url: str, category: str) -> bool:
    if category == "Claude/Anthropic":
        return True
    return any(kw in url.lower() for kw in CLAUDE_KEYWORDS)


def get_relevance_score(category: str, url: str) -> int:
    low_relevance_patterns = ["bill", "recharge", "porter", "referral", "photos.google"]
    if any(p in url.lower() for p in low_relevance_patterns):
        return 1
    return RELEVANCE_SCORES.get(category, 2)


def get_claude_accuracy_note(url: str) -> str:
    url_lower = url.lower()
    for keyword, note in CLAUDE_ACCURACY_NOTES.items():
        if keyword == "default":
            continue
        if keyword.replace(" ", "") in url_lower.replace("-", "").replace("_", ""):
            return note
    return CLAUDE_ACCURACY_NOTES["default"]


# ---------------------------------------------------------------------------
# Excel Generation
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
CLAUDE_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
BODY_FONT = Font(name="Arial", size=10)
LINK_FONT = Font(name="Arial", size=10, color="0563C1", underline="single")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="top")


def style_header(ws, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def build_all_links_sheet(ws, data):
    headers = ["S.No.", "Date", "Link", "Topic/Category", "Claude Related", "Relevance (1-5)"]
    ws.append(headers)
    style_header(ws, len(headers))
    for i, item in enumerate(data, 1):
        row = i + 1
        ws.cell(row=row, column=1, value=i).font = BODY_FONT
        ws.cell(row=row, column=1).alignment = CENTER_ALIGNMENT
        ws.cell(row=row, column=2, value=item["date"]).font = BODY_FONT
        ws.cell(row=row, column=2).alignment = CENTER_ALIGNMENT
        cell = ws.cell(row=row, column=3, value=item["url"])
        cell.font = LINK_FONT
        try:
            cell.hyperlink = item["url"]
        except Exception:
            pass
        ws.cell(row=row, column=4, value=item["category"]).font = BODY_FONT
        ws.cell(row=row, column=5, value="Yes" if item["claude"] else "No").font = BODY_FONT
        ws.cell(row=row, column=5).alignment = CENTER_ALIGNMENT
        ws.cell(row=row, column=6, value=item["relevance"]).font = BODY_FONT
        ws.cell(row=row, column=6).alignment = CENTER_ALIGNMENT
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = THIN_BORDER
        if item["claude"]:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = CLAUDE_FILL
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 70
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F{len(data) + 1}"


def build_claude_sheet(ws, claude_links):
    headers = ["S.No.", "Date", "Link", "Sub-Topic", "Accuracy Note"]
    ws.append(headers)
    style_header(ws, len(headers))
    for i, item in enumerate(claude_links, 1):
        row = i + 1
        ws.cell(row=row, column=1, value=i).font = BODY_FONT
        ws.cell(row=row, column=1).alignment = CENTER_ALIGNMENT
        ws.cell(row=row, column=2, value=item["date"]).font = BODY_FONT
        ws.cell(row=row, column=2).alignment = CENTER_ALIGNMENT
        cell = ws.cell(row=row, column=3, value=item["url"])
        cell.font = LINK_FONT
        try:
            cell.hyperlink = item["url"]
        except Exception:
            pass
        ws.cell(row=row, column=4, value=item["category"]).font = BODY_FONT
        ws.cell(row=row, column=5, value=item["accuracy_note"]).font = BODY_FONT
        ws.cell(row=row, column=5).alignment = WRAP_ALIGNMENT
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = THIN_BORDER
            ws.cell(row=row, column=col).fill = CLAUDE_FILL
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 70
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 55
    ws.freeze_panes = "A2"


def build_summary_sheet(ws, data):
    dates = [d["date"] for d in data if d["date"] != "Unknown"]
    categories = Counter(d["category"] for d in data)
    claude_count = sum(1 for d in data if d["claude"])
    unique_dates = len(set(dates))
    stats = [
        ("Metric", "Value"),
        ("Total Links", len(data)),
        ("Date Range", f"{min(dates) if dates else 'N/A'} to {max(dates) if dates else 'N/A'}"),
        ("Unique Dates", unique_dates),
        ("Avg Links/Day", round(len(data) / max(unique_dates, 1), 1)),
        ("Claude-Related Links", claude_count),
        ("Claude % of Total", f"{claude_count / max(len(data), 1) * 100:.1f}%"),
        ("", ""),
        ("Top Categories", "Count"),
    ]
    for cat, count in categories.most_common(15):
        stats.append((cat, count))
    for row_idx, (metric, value) in enumerate(stats, 1):
        ws.cell(row=row_idx, column=1, value=metric).font = Font(name="Arial", bold=(row_idx in (1, 9)), size=11)
        ws.cell(row=row_idx, column=2, value=value).font = Font(name="Arial", bold=(row_idx in (1, 9)), size=11)
        ws.cell(row=row_idx, column=1).border = THIN_BORDER
        ws.cell(row=row_idx, column=2).border = THIN_BORDER
    for col in (1, 2):
        ws.cell(row=1, column=col).fill = HEADER_FILL
        ws.cell(row=1, column=col).font = HEADER_FONT
        ws.cell(row=9, column=col).fill = HEADER_FILL
        ws.cell(row=9, column=col).font = HEADER_FONT
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 30


def generate_excel(data, output_path):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "All Links"
    build_all_links_sheet(ws1, data)
    ws2 = wb.create_sheet("Claude Topics")
    build_claude_sheet(ws2, [d for d in data if d["claude"]])
    ws3 = wb.create_sheet("Summary")
    build_summary_sheet(ws3, data)
    wb.save(output_path)
    return output_path


# ---------------------------------------------------------------------------
# Pipeline
# ---------------------------------------------------------------------------

def analyze(text: str) -> list[dict]:
    raw_links = extract_links(text)
    results = []
    for item in raw_links:
        url = item["url"]
        category = categorize_url(url)
        claude = is_claude_related(url, category)
        relevance = get_relevance_score(category, url)
        accuracy_note = get_claude_accuracy_note(url) if claude else ""
        results.append({
            "date": item["date"], "url": url, "category": category,
            "claude": claude, "relevance": relevance, "accuracy_note": accuracy_note,
        })
    return results


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    input_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else "output.xlsx"
    if input_path == "-":
        text = sys.stdin.read()
    else:
        if not os.path.isfile(input_path):
            print(f"ERROR: File not found: {input_path}")
            sys.exit(1)
        with open(input_path, "r", encoding="utf-8", errors="replace") as f:
            text = f.read()
    if not text.strip():
        print("ERROR: Input is empty.")
        sys.exit(1)
    links_data = analyze(text)
    if not links_data:
        print("WARNING: No URLs found in the input.")
        sys.exit(0)
    generate_excel(links_data, output_path)
    claude_count = sum(1 for d in links_data if d["claude"])
    cats = Counter(d["category"] for d in links_data)
    print(f"✅ Analyzed {len(links_data)} links → {output_path}")
    print(f"   Claude-related: {claude_count} ({claude_count/len(links_data)*100:.1f}%)")
    print(f"   Categories: {len(cats)}")
    print(f"   Top 5: {', '.join(f'{c} ({n})' for c, n in cats.most_common(5))}")


if __name__ == "__main__":
    main()
