#!/usr/bin/env python3
"""
Link Analyzer Agent — AI-powered WhatsApp link analysis using Claude API.

Uses Claude to intelligently categorize, score, and verify links
instead of dumb keyword matching.

Usage:
    python analyze.py <chat_file.txt> [output.xlsx]
    cat chat.txt | python analyze.py - [output.xlsx]

Requires:
    ANTHROPIC_API_KEY — set via environment variable OR .env file
    pip install anthropic openpyxl python-dotenv
"""

import re
import sys
import os
import json
import time
from datetime import datetime
from collections import Counter
from urllib.parse import urlparse

# Load .env file if present (so user can put API key there)
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass  # dotenv is optional — user can set env var directly instead

try:
    import anthropic
except ImportError:
    print("ERROR: anthropic SDK required. Install with: pip install anthropic")
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("ERROR: openpyxl required. Install with: pip install openpyxl")
    sys.exit(1)


# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

MODEL = "claude-sonnet-4-20250514"
BATCH_SIZE = 30  # links per API call (balances cost vs. context)
MAX_RETRIES = 3

CATEGORIES = [
    "Claude/Anthropic", "Voice AI / TTS", "AI Prompting", "RAG / Retrieval",
    "Coding / Dev Tools", "AI Agents", "Education / Learning",
    "Product Management", "Business / Sales", "Career / Jobs",
    "Tech Companies", "Open Source", "Health / Wellness", "Finance",
    "India / Culture", "AI / ML General", "Robotics", "YouTube",
    "Facebook", "Instagram", "Twitter/X", "LinkedIn Post - General", "Other",
]

SYSTEM_PROMPT = """You are a link categorization expert. You analyze URLs shared in WhatsApp groups and categorize them accurately.

You will receive a batch of URLs with dates. For EACH URL, analyze the URL structure, path, slugs, and any embedded text to determine:

1. **category**: Pick the BEST fit from this list:
   Claude/Anthropic, Voice AI / TTS, AI Prompting, RAG / Retrieval, Coding / Dev Tools, AI Agents, Education / Learning, Product Management, Business / Sales, Career / Jobs, Tech Companies, Open Source, Health / Wellness, Finance, India / Culture, AI / ML General, Robotics, YouTube, Facebook, Instagram, Twitter/X, LinkedIn Post - General, Other

2. **is_claude_related**: true/false — Is this about Claude, Anthropic, or related tools?
   - Official Anthropic: claude, anthropic, cowork, claude code, sonnet, opus, haiku → true
   - Community tools built for Claude: openclaw, clawdbot, zeroclaw, nanoclaw, picoclaw → true
   - Third-party Claude tools: moltbot, moltbook → true
   - Generic AI that mentions Claude in passing → false

3. **relevance**: Score 1-5
   - 5: Claude/Anthropic content
   - 4: AI/ML, Voice AI, Coding, RAG, Education, Agents
   - 3: Business, Career, Tech Companies, Finance
   - 2: Social media (YouTube, Facebook), Culture
   - 1: Utility bills, referral links, personal photos

4. **accuracy_note**: ONLY for Claude-related links. Include:
   - Whether it's an official Anthropic product or community/third-party
   - What to verify (model versions, feature claims, etc.)
   - Empty string "" for non-Claude links

5. **topic_summary**: A brief 5-10 word description of what the link is about, inferred from the URL.

RESPOND WITH ONLY valid JSON array. No markdown, no backticks, no explanation.
Example:
[
  {
    "index": 0,
    "category": "Claude/Anthropic",
    "is_claude_related": true,
    "relevance": 5,
    "accuracy_note": "Official Anthropic CLI tool. Verify version claims against docs.anthropic.com.",
    "topic_summary": "Claude Code tips and workflow tricks"
  }
]"""


# ---------------------------------------------------------------------------
# Parsing (deterministic — no AI needed here)
# ---------------------------------------------------------------------------

DATE_PATTERNS = [
    r"(\d{1,2}/\d{1,2}/\d{2,4}),?\s*\d{1,2}:\d{2}",
    r"(\d{4}-\d{2}-\d{2})\s*\d{1,2}:\d{2}",
    r"(\d{1,2}/\d{1,2}/\d{4})",
    r"\[(\d{1,2}/\d{1,2}/\d{2,4}),?\s*\d{1,2}:\d{2}:\d{2}\]",
]

URL_PATTERN = re.compile(r'https?://[^\s<>"\')\]]+')


def parse_date(date_str: str) -> str:
    """Normalize a date string to YYYY-MM-DD format."""
    date_str = date_str.strip().strip("[]")
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%m/%d/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(date_str.rstrip(","), fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return date_str


def extract_links(text: str) -> list[dict]:
    """Extract all URLs with dates from chat text."""
    results = []
    current_date = None

    for line in text.splitlines():
        for pattern in DATE_PATTERNS:
            match = re.search(pattern, line)
            if match:
                current_date = parse_date(match.group(1))
                break

        for url in URL_PATTERN.findall(line):
            url = url.rstrip(".,;:!?)]}>")
            results.append({"date": current_date or "Unknown", "url": url})

    return results


# ---------------------------------------------------------------------------
# AI Analysis (the actual agent part)
# ---------------------------------------------------------------------------

def analyze_batch_with_claude(client: anthropic.Anthropic, links_batch: list[dict], batch_num: int, total_batches: int) -> list[dict]:
    """Send a batch of links to Claude for intelligent categorization."""

    # Build the prompt with indexed URLs
    url_list = "\n".join(
        f"[{i}] Date: {link['date']} | URL: {link['url']}"
        for i, link in enumerate(links_batch)
    )

    user_prompt = f"""Analyze these {len(links_batch)} URLs and categorize each one.
Return a JSON array with one object per URL, matching the index.

URLs:
{url_list}"""

    for attempt in range(MAX_RETRIES):
        try:
            response = client.messages.create(
                model=MODEL,
                max_tokens=4096,
                system=SYSTEM_PROMPT,
                messages=[{"role": "user", "content": user_prompt}],
            )

            # Extract text response
            text = response.content[0].text.strip()

            # Clean potential markdown wrapping
            if text.startswith("```"):
                text = re.sub(r"^```(?:json)?\s*", "", text)
                text = re.sub(r"\s*```$", "", text)

            results = json.loads(text)

            # Validate we got results for all links
            if len(results) != len(links_batch):
                print(f"  ⚠ Batch {batch_num}: Expected {len(links_batch)} results, got {len(results)}. Padding missing entries.")
                # Pad missing entries with defaults
                result_by_index = {r["index"]: r for r in results}
                padded = []
                for i in range(len(links_batch)):
                    if i in result_by_index:
                        padded.append(result_by_index[i])
                    else:
                        padded.append({
                            "index": i,
                            "category": "Other",
                            "is_claude_related": False,
                            "relevance": 2,
                            "accuracy_note": "",
                            "topic_summary": "Unable to categorize",
                        })
                results = padded

            return results

        except json.JSONDecodeError as e:
            print(f"  ⚠ Batch {batch_num}, attempt {attempt+1}: JSON parse error — {e}")
            if attempt < MAX_RETRIES - 1:
                time.sleep(2 ** attempt)
            continue
        except anthropic.RateLimitError:
            wait = 2 ** (attempt + 2)
            print(f"  ⏳ Rate limited. Waiting {wait}s...")
            time.sleep(wait)
            continue
        except anthropic.APIError as e:
            print(f"  ❌ API error: {e}")
            if attempt < MAX_RETRIES - 1:
                time.sleep(2 ** attempt)
            continue

    # All retries failed — return defaults
    print(f"  ❌ Batch {batch_num}: All retries failed. Using fallback categorization.")
    return [
        {
            "index": i,
            "category": "Other",
            "is_claude_related": False,
            "relevance": 2,
            "accuracy_note": "",
            "topic_summary": "Categorization failed",
        }
        for i in range(len(links_batch))
    ]


def analyze_with_ai(links: list[dict]) -> list[dict]:
    """Main AI analysis pipeline — batches links through Claude."""

    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        print("ERROR: ANTHROPIC_API_KEY not found.")
        print()
        print("  Option 1 — .env file (recommended):")
        print("    Create a .env file in this directory with:")
        print("    ANTHROPIC_API_KEY=sk-ant-your-key-here")
        print()
        print("  Option 2 — Environment variable:")
        print("    export ANTHROPIC_API_KEY='sk-ant-your-key-here'")
        print()
        print("  Get your key at: https://console.anthropic.com/settings/keys")
        sys.exit(1)

    client = anthropic.Anthropic(api_key=api_key)

    # Split into batches
    batches = [links[i:i + BATCH_SIZE] for i in range(0, len(links), BATCH_SIZE)]
    total_batches = len(batches)

    print(f"🤖 Analyzing {len(links)} links in {total_batches} batches using {MODEL}...")

    all_results = []

    for batch_idx, batch in enumerate(batches, 1):
        print(f"  📦 Batch {batch_idx}/{total_batches} ({len(batch)} links)...", end=" ", flush=True)

        ai_results = analyze_batch_with_claude(client, batch, batch_idx, total_batches)

        # Merge AI analysis with original link data
        for link, ai_result in zip(batch, ai_results):
            all_results.append({
                "date": link["date"],
                "url": link["url"],
                "category": ai_result.get("category", "Other"),
                "claude": ai_result.get("is_claude_related", False),
                "relevance": ai_result.get("relevance", 2),
                "accuracy_note": ai_result.get("accuracy_note", ""),
                "topic_summary": ai_result.get("topic_summary", ""),
            })

        print("✓")

        # Small delay between batches to avoid rate limits
        if batch_idx < total_batches:
            time.sleep(1)

    return all_results


# ---------------------------------------------------------------------------
# Excel Generation
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
CLAUDE_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
BODY_FONT = Font(name="Arial", size=10)
LINK_FONT = Font(name="Arial", size=10, color="0563C1", underline="single")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="top")


def style_header(ws, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def build_all_links_sheet(ws, data):
    headers = ["S.No.", "Date", "Link", "Topic/Category", "Topic Summary", "Claude Related", "Relevance (1-5)"]
    ws.append(headers)
    style_header(ws, len(headers))

    for i, item in enumerate(data, 1):
        row = i + 1
        ws.cell(row=row, column=1, value=i).font = BODY_FONT
        ws.cell(row=row, column=1).alignment = CENTER
        ws.cell(row=row, column=2, value=item["date"]).font = BODY_FONT
        ws.cell(row=row, column=2).alignment = CENTER

        cell = ws.cell(row=row, column=3, value=item["url"])
        cell.font = LINK_FONT
        try:
            cell.hyperlink = item["url"]
        except Exception:
            pass

        ws.cell(row=row, column=4, value=item["category"]).font = BODY_FONT
        ws.cell(row=row, column=5, value=item.get("topic_summary", "")).font = BODY_FONT
        ws.cell(row=row, column=5).alignment = WRAP
        ws.cell(row=row, column=6, value="Yes" if item["claude"] else "No").font = BODY_FONT
        ws.cell(row=row, column=6).alignment = CENTER
        ws.cell(row=row, column=7, value=item["relevance"]).font = BODY_FONT
        ws.cell(row=row, column=7).alignment = CENTER

        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = THIN_BORDER
        if item["claude"]:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = CLAUDE_FILL

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 65
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 35
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 16
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{len(data) + 1}"


def build_claude_sheet(ws, claude_links):
    headers = ["S.No.", "Date", "Link", "Topic Summary", "Accuracy Note"]
    ws.append(headers)
    style_header(ws, len(headers))

    for i, item in enumerate(claude_links, 1):
        row = i + 1
        ws.cell(row=row, column=1, value=i).font = BODY_FONT
        ws.cell(row=row, column=1).alignment = CENTER
        ws.cell(row=row, column=2, value=item["date"]).font = BODY_FONT
        ws.cell(row=row, column=2).alignment = CENTER

        cell = ws.cell(row=row, column=3, value=item["url"])
        cell.font = LINK_FONT
        try:
            cell.hyperlink = item["url"]
        except Exception:
            pass

        ws.cell(row=row, column=4, value=item.get("topic_summary", "")).font = BODY_FONT
        ws.cell(row=row, column=4).alignment = WRAP
        ws.cell(row=row, column=5, value=item["accuracy_note"]).font = BODY_FONT
        ws.cell(row=row, column=5).alignment = WRAP

        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = THIN_BORDER
            ws.cell(row=row, column=col).fill = CLAUDE_FILL

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 65
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 55
    ws.freeze_panes = "A2"


def build_summary_sheet(ws, data):
    dates = [d["date"] for d in data if d["date"] != "Unknown"]
    categories = Counter(d["category"] for d in data)
    claude_count = sum(1 for d in data if d["claude"])
    unique_dates = len(set(dates))

    stats = [
        ("Metric", "Value"),
        ("Total Links", len(data)),
        ("Date Range", f"{min(dates) if dates else 'N/A'} to {max(dates) if dates else 'N/A'}"),
        ("Unique Dates", unique_dates),
        ("Avg Links/Day", round(len(data) / max(unique_dates, 1), 1)),
        ("Claude-Related Links", claude_count),
        ("Claude % of Total", f"{claude_count / max(len(data), 1) * 100:.1f}%"),
        ("AI Model Used", MODEL),
        ("", ""),
        ("Top Categories", "Count"),
    ]
    for cat, count in categories.most_common(15):
        stats.append((cat, count))

    for row_idx, (metric, value) in enumerate(stats, 1):
        ws.cell(row=row_idx, column=1, value=metric).font = Font(name="Arial", bold=(row_idx in (1, 10)), size=11)
        ws.cell(row=row_idx, column=2, value=value).font = Font(name="Arial", bold=(row_idx in (1, 10)), size=11)
        ws.cell(row=row_idx, column=1).border = THIN_BORDER
        ws.cell(row=row_idx, column=2).border = THIN_BORDER

    for col in (1, 2):
        ws.cell(row=1, column=col).fill = HEADER_FILL
        ws.cell(row=1, column=col).font = HEADER_FONT
        ws.cell(row=10, column=col).fill = HEADER_FILL
        ws.cell(row=10, column=col).font = HEADER_FONT

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 30


def generate_excel(data, output_path):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "All Links"
    build_all_links_sheet(ws1, data)

    ws2 = wb.create_sheet("Claude Topics")
    build_claude_sheet(ws2, [d for d in data if d["claude"]])

    ws3 = wb.create_sheet("Summary")
    build_summary_sheet(ws3, data)

    wb.save(output_path)
    return output_path


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    input_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else "output.xlsx"

    # Read input
    if input_path == "-":
        text = sys.stdin.read()
    else:
        if not os.path.isfile(input_path):
            print(f"ERROR: File not found: {input_path}")
            sys.exit(1)
        with open(input_path, "r", encoding="utf-8", errors="replace") as f:
            text = f.read()

    if not text.strip():
        print("ERROR: Input is empty.")
        sys.exit(1)

    # Step 1: Parse (deterministic)
    links = extract_links(text)
    if not links:
        print("WARNING: No URLs found in the input.")
        sys.exit(0)
    print(f"📎 Found {len(links)} links in the chat export.")

    # Step 2: AI Analysis
    analyzed = analyze_with_ai(links)

    # Step 3: Generate Excel
    generate_excel(analyzed, output_path)

    # Summary
    claude_count = sum(1 for d in analyzed if d["claude"])
    cats = Counter(d["category"] for d in analyzed)
    print(f"\n✅ Done → {output_path}")
    print(f"   Claude-related: {claude_count} ({claude_count/len(analyzed)*100:.1f}%)")
    print(f"   Categories: {len(cats)}")
    print(f"   Top 5: {', '.join(f'{c} ({n})' for c, n in cats.most_common(5))}")


if __name__ == "__main__":
    main()
